"""
Excel processor for FL DOE Standards benchmarks.
Handles reading and processing of benchmark definitions from Excel source.
Provides functionality to save/load processed benchmarks via pickle.
"""

import pandas as pd
import pickle
import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List
import logging
import re
from dataclasses import dataclass

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@dataclass
class Benchmark:
    """Represents a benchmark definition and metadata."""
    id: str
    definition: str
    grade_level: str
    subject: str = "Mathematics"
    cpalms_url: str = ""

class ExcelProcessingError(Exception):
    """Custom exception for Excel processing errors."""
    pass

def get_cell_hyperlink(sheet, row, col) -> Optional[str]:
    """
    Get the hyperlink URL from a specific cell, handling merged cells.
    
    Args:
        sheet: Excel worksheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        
    Returns:
        Hyperlink URL or None if no hyperlink exists
    """
    try:
        # Check if the cell is part of a merged range
        for merged_range in sheet.merged_cells.ranges:
            if row >= merged_range.min_row and row <= merged_range.max_row and \
               col >= merged_range.min_col and col <= merged_range.max_col:
                # Use the top-left cell of the merged range
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                if hasattr(cell, 'hyperlink') and cell.hyperlink:
                    logger.info(f"Found hyperlink in merged cell range {merged_range}, using cell ({merged_range.min_row}, {merged_range.min_col})")
                    return cell.hyperlink.target
                return None
        
        # If not in a merged range, check the cell directly
        cell = sheet.cell(row=row, column=col)
        if hasattr(cell, 'hyperlink') and cell.hyperlink:
            return cell.hyperlink.target
        return None
    except Exception as e:
        logger.error(f"Error getting hyperlink from cell ({row}, {col}): {e}")
        return None

def is_valid_url(url: str) -> bool:
    """
    Check if a string looks like a valid URL.
    
    Args:
        url: String to check
        
    Returns:
        True if the string looks like a URL, False otherwise
    """
    if not url:
        return False
    
    # Check if it starts with http:// or https://
    if url.startswith(('http://', 'https://')):
        return True
    
    # Check if it looks like a URL (contains domain-like pattern)
    url_pattern = re.compile(r'^(www\.)?[a-zA-Z0-9][-a-zA-Z0-9.]*\.[a-zA-Z]{2,}(/.*)?$')
    return bool(url_pattern.match(url))

def process_excel_benchmarks(file_path: str) -> Dict[str, Benchmark]:
    """
    Process BEST Math Extract Excel file into benchmark dictionary.
    
    Args:
        file_path: Path to the Excel file containing benchmark definitions
        
    Returns:
        Dictionary mapping benchmark IDs to Benchmark objects
        
    Raises:
        ExcelProcessingError: If there's an error reading or processing the Excel file
        FileNotFoundError: If the Excel file doesn't exist
    """
    try:
        # Convert string path to Path object
        excel_path = Path(file_path)
        
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
            
        logger.info(f"Reading Excel file: {file_path}")
        
        # Load workbook with openpyxl to access hyperlinks
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        sheet = wb.active
        
        # Skip the first two rows and use the third row as column names
        df = pd.read_excel(excel_path, skiprows=2)
        
        # Find the column index for "Direct Link"
        direct_link_col = None
        for i, col_name in enumerate(df.columns):
            if col_name == "Direct Link":
                direct_link_col = i + 1  # +1 because openpyxl is 1-indexed
                logger.info(f"Found 'Direct Link' column at index {direct_link_col}")
                break
                
        if direct_link_col is None:
            logger.warning("Could not find 'Direct Link' column in Excel file")
        
        # Initialize dictionary for benchmarks
        benchmarks: Dict[str, Benchmark] = {}
        
        # Create a mapping of benchmark IDs to all their related rows
        benchmark_rows = {}
        current_benchmark = None
        
        # First pass: collect all rows for each benchmark ID, including rows without benchmark ID
        for idx, row in df.iterrows():
            try:
                benchmark_id = row['Benchmark#']
                
                if pd.notna(benchmark_id):
                    # New benchmark found
                    current_benchmark = str(benchmark_id).strip()
                    if current_benchmark not in benchmark_rows:
                        benchmark_rows[current_benchmark] = []
                    benchmark_rows[current_benchmark].append(idx)
                elif current_benchmark is not None:
                    # This row belongs to the current benchmark
                    benchmark_rows[current_benchmark].append(idx)
            except Exception as e:
                logger.error(f"Error collecting rows for benchmark at row {idx}: {e}")
        
        logger.info(f"Found {len(benchmark_rows)} unique benchmarks")
        
        # Create a mapping of benchmark IDs to their URLs
        benchmark_urls = {}
        
        # Second pass: find URLs for each benchmark by checking all its rows
        for benchmark_id, row_indices in benchmark_rows.items():
            try:
                # Check all rows for this benchmark for a valid URL
                for idx in row_indices:
                    row = df.iloc[idx]
                    
                    # Calculate the actual row number in the Excel file (add 3 for header rows)
                    excel_row = idx + 3
                    
                    # Get hyperlink directly from the cell, handling merged cells
                    url = get_cell_hyperlink(sheet, excel_row, direct_link_col)
                    
                    if url and is_valid_url(url):
                        logger.info(f"Found valid URL for benchmark {benchmark_id} at row {excel_row}: {url}")
                        benchmark_urls[benchmark_id] = url
                        break  # Use the first valid URL found
                    elif 'Direct Link' in row and pd.notna(row['Direct Link']):
                        # If we have text but no hyperlink, log it
                        cell_text = str(row['Direct Link']).strip()
                        logger.info(f"Found text but no hyperlink for benchmark {benchmark_id} at row {excel_row}: {cell_text}")
                
                # If no valid URL was found after checking all rows
                if benchmark_id not in benchmark_urls:
                    logger.info(f"No valid URL found for benchmark {benchmark_id} after checking {len(row_indices)} rows")
            except Exception as e:
                logger.error(f"Error processing URL for benchmark {benchmark_id}: {e}")
        
        logger.info(f"Found URLs for {len(benchmark_urls)} benchmarks")
        
        # Third pass: process benchmarks and use the collected URLs
        current_benchmark = None
        current_description = []
        current_grade = None
        
        for idx, row in df.iterrows():
            try:
                # Check if this row has a benchmark ID
                benchmark_id = row['Benchmark#']
                
                # If this is a new benchmark and we have a previous one, save it
                if pd.notna(benchmark_id) and current_benchmark is not None:
                    # Get the URL for this benchmark
                    cpalms_url = benchmark_urls.get(current_benchmark, "")
                    
                    benchmarks[current_benchmark] = Benchmark(
                        id=current_benchmark,
                        definition=' '.join(current_description).strip(),
                        grade_level=current_grade or "Unknown",
                        cpalms_url=cpalms_url
                    )
                    # Reset for the new benchmark
                    current_description = []
                
                # If this is a benchmark row, update the current benchmark
                if pd.notna(benchmark_id):
                    current_benchmark = benchmark_id.strip()
                    current_grade = row['Grade'] if pd.notna(row['Grade']) else "Unknown"
                
                # Add description text if it exists
                if pd.notna(row['Description']):
                    current_description.append(str(row['Description']).strip())
                    
            except KeyError as e:
                logger.error(f"Missing required column in row {idx}: {e}")
                raise ExcelProcessingError(f"Excel format error: Missing column {e}")
            except Exception as e:
                logger.error(f"Error processing row {idx}: {e}")
                continue
        
        # Save the last benchmark if there is one
        if current_benchmark is not None and current_description:
            # Get the URL for this benchmark
            cpalms_url = benchmark_urls.get(current_benchmark, "")
                    
            benchmarks[current_benchmark] = Benchmark(
                id=current_benchmark,
                definition=' '.join(current_description).strip(),
                grade_level=current_grade or "Unknown",
                cpalms_url=cpalms_url
            )
                
        logger.info(f"Successfully processed {len(benchmarks)} benchmarks")
        return benchmarks
        
    except pd.errors.EmptyDataError:
        logger.error("Excel file is empty")
        raise ExcelProcessingError("Excel file contains no data")
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        raise ExcelProcessingError(f"Failed to process Excel file: {str(e)}")

def get_benchmark(benchmark_id: str, benchmarks: Dict[str, Benchmark]) -> Optional[Benchmark]:
    """
    Retrieve a benchmark by ID.
    
    Args:
        benchmark_id: The ID of the benchmark to retrieve
        benchmarks: Dictionary of benchmarks
        
    Returns:
        Benchmark object if found, None otherwise
    """
    return benchmarks.get(benchmark_id)

def save_benchmarks_pickle(benchmarks: Dict[str, Benchmark], file_path: str = "data/processed/benchmarks.pkl"):
    """
    Save processed benchmarks to pickle file.
    
    Args:
        benchmarks: Dictionary of benchmark objects
        file_path: Path to save pickle file
        
    Raises:
        IOError: If unable to write to file
    """
    try:
        with open(file_path, 'wb') as f:
            pickle.dump(benchmarks, f)
        logger.info(f"Saved {len(benchmarks)} benchmarks to {file_path}")
    except Exception as e:
        logger.error(f"Failed to save benchmarks pickle: {e}")
        raise

def load_benchmarks_pickle(file_path: str = "data/processed/benchmarks.pkl") -> Dict[str, Benchmark]:
    """
    Load benchmarks from pickle file.
    
    Args:
        file_path: Path to pickle file
        
    Returns:
        Dictionary of benchmark objects
        
    Raises:
        FileNotFoundError: If pickle file doesn't exist
        IOError: If unable to read file
    """
    try:
        with open(file_path, 'rb') as f:
            benchmarks = pickle.load(f)
        logger.info(f"Loaded {len(benchmarks)} benchmarks from {file_path}")
        return benchmarks
    except Exception as e:
        logger.error(f"Failed to load benchmarks pickle: {e}")
        raise

if __name__ == "__main__":
    try:
        # Process Excel file
        excel_path = "data/raw/BEST Math Extract.xlsx"
        benchmarks = process_excel_benchmarks(excel_path)
        logger.info(f"Total benchmarks processed: {len(benchmarks)}")
        
        # Save to pickle file
        save_benchmarks_pickle(benchmarks)
    except Exception as e:
        logger.error(f"Failed to process benchmarks: {e}")
